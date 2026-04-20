"""
단국테크 사원DB 계정정보 Excel → SQLite 가져오기 스크립트
사용법: python import_users.py
"""

import sys
from pathlib import Path

# openpyxl 없으면 안내
try:
    import openpyxl
except ImportError:
    print("openpyxl 패키지가 필요합니다: pip install openpyxl")
    sys.exit(1)

from auth import DB_PATH, _get_db, _hash, init_db
import secrets

EXCEL_PATH = Path("../data/docs/단국테크_사원DB_계정정보.xlsx")

# 권한 레벨 → role 매핑
ROLE_MAP = {
    "SUPER_ADMIN": "admin",
    "ADMIN":       "admin",
    "USER":        "user",
}


def import_from_excel(excel_path: Path, clear_existing: bool = True):
    if not excel_path.exists():
        print(f"파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["계정관리대장"]

    # 헤더 행 찾기 (사번 컬럼이 있는 행)
    header_row = None
    for row in ws.iter_rows(values_only=True):
        if row[0] == "사번":
            header_row = row
            break

    if not header_row:
        print("헤더 행을 찾을 수 없습니다.")
        sys.exit(1)

    # 컬럼 인덱스 매핑
    col = {name: i for i, name in enumerate(header_row)}

    # DB 초기화
    init_db()
    conn = _get_db()

    if clear_existing:
        conn.execute("DELETE FROM users")
        conn.commit()
        print("기존 계정 데이터를 초기화했습니다.")

    inserted = 0
    skipped  = 0

    for row in ws.iter_rows(values_only=True):
        # 헤더나 빈 행, 타이틀 행 스킵
        if not row[col["사번"]] or row[col["사번"]] == "사번":
            continue
        # 사번 형식이 아닌 경우 스킵
        employee_id = str(row[col["사번"]]).strip()
        if not employee_id.startswith("DK"):
            continue

        username      = str(row[col["로그인 ID"]]).strip()
        password      = str(row[col["초기 비밀번호"]]).strip()
        level         = str(row[col["권한 레벨"]]).strip()
        name          = str(row[col["성명"]]).strip()
        department    = str(row[col["부서"]]).strip()
        position      = str(row[col["직급"]]).strip()
        email         = str(row[col["이메일"]]).strip()
        status        = str(row[col["계정 상태"]]).strip()
        role          = ROLE_MAP.get(level, "user")

        salt          = secrets.token_hex(16)
        password_hash = _hash(password, salt)

        try:
            conn.execute(
                """INSERT OR REPLACE INTO users
                   (employee_id, username, password_hash, salt, role,
                    name, department, position, email, status)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (employee_id, username, password_hash, salt, role,
                 name, department, position, email, status),
            )
            inserted += 1
        except Exception as e:
            print(f"  ⚠ 스킵 ({username}): {e}")
            skipped += 1

    conn.commit()
    conn.close()

    print(f"\n임포트 완료: {inserted}명 등록, {skipped}명 스킵")
    print(f"DB 경로: {DB_PATH.resolve()}")


def print_summary():
    conn = _get_db()
    rows = conn.execute(
        "SELECT role, COUNT(*) as cnt FROM users GROUP BY role"
    ).fetchall()
    conn.close()
    print("\n[권한별 계정 수]")
    for r in rows:
        label = "관리자(admin)" if r["role"] == "admin" else "일반사용자(user)"
        print(f"  {label}: {r['cnt']}명")


if __name__ == "__main__":
    print(f"Excel 파일 읽는 중: {EXCEL_PATH.resolve()}")
    import_from_excel(EXCEL_PATH)
    print_summary()
